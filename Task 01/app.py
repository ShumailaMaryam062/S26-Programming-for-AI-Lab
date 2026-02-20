from flask import Flask, render_template, request, jsonify, send_file
import requests
from requests.exceptions import MissingSchema, RequestException
from modules.scrapper import Scrapper
from modules.info_reader import InfoReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
import os
from datetime import datetime

app = Flask(__name__)

# Create output directory if it doesn't exist
if not os.path.exists('output'):
    os.makedirs('output')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/scrape', methods=['POST'])
def scrape():
    try:
        data = request.get_json()
        url = data.get('url')
        
        if not url:
            return jsonify({'error': 'URL is required'}), 400
        
        # Ensure URL has a scheme
        if not url.startswith('http://') and not url.startswith('https://'):
            url = 'http://' + url
        
        # Verify the URL is accessible
        try:
            response = requests.get(url, timeout=5, verify=False)
        except RequestException as e:
            return jsonify({'error': f'Could not reach URL: {str(e)}'}), 400
        
        # Run the scraper
        scrapper = Scrapper(url=url)
        info_data = scrapper.getText()
        reader = InfoReader(content=info_data)
        
        return jsonify({
            'success': True,
            'url': url,
            'emails': reader.getEmails()
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/scrape-multiple', methods=['POST'])
def scrape_multiple():
    try:
        data = request.get_json()
        urls = data.get('urls', [])
        
        if not urls:
            return jsonify({'error': 'URLs are required'}), 400
        
        results = []
        for url in urls:
            if not url:
                continue
                
            if not url.startswith('http://') and not url.startswith('https://'):
                url = 'http://' + url
            
            try:
                response = requests.get(url, timeout=5, verify=False)
                scrapper = Scrapper(url=url)
                info_data = scrapper.getText()
                reader = InfoReader(content=info_data)
                
                results.append({
                    'url': url,
                    'success': True,
                    'emails': reader.getEmails()
                })
            except Exception as e:
                results.append({
                    'url': url,
                    'success': False,
                    'error': str(e)
                })
        
        return jsonify({'results': results})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({'error': 'No results to export'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"
        
        # Add headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = 'URL'
        ws['B1'] = 'Email'
        
        for col in ['A1', 'B1']:
            ws[col].fill = header_fill
            ws[col].font = header_font
            ws[col].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        row = 2
        for result in results:
            if result.get('success'):
                url = result.get('url', '')
                emails = result.get('emails', [])
                
                # Add all emails in one row, separated by semicolons
                ws[f'A{row}'] = url
                ws[f'B{row}'] = '; '.join(emails) if emails else ''
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        
        # Save file
        filename = f"output/scraper_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return jsonify({
            'success': True,
            'filename': filename
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<path:filename>', methods=['GET'])
def download_file(filename):
    try:
        return send_file(filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=20002)
